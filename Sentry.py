import re
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import PyPDF2
from datetime import datetime
import os
import logging
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# Para gráficos
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib
matplotlib.use('TkAgg')

# Configuração do logging
logging.basicConfig(
    filename='cnd_dashboard.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    encoding='utf-8'
)


class CNDDashboard:
    def __init__(self):
        ctk.set_appearance_mode("dark")
        ctk.set_default_color_theme("blue")

        self.root = ctk.CTk()
        self.root.title("Sentry - Controle de Certidões")
        self.root.geometry("1400x800")
        self.root.minsize(1200, 700)

        self.folder_path = tk.StringVar()
        self.processing = False
        self.cancel_requested = False
        self.executor = None
        self.is_closing = False
        self.results_data = []
        self.filtered_data = []
        self.search_var = tk.StringVar()
        self.sort_column = None
        self.sort_reverse = False

        # Contadores para o dashboard
        self.stats = {
            "total": 0,
            "completo": 0,
            "incompleto": 0,
            "vencidas": 0,
            "validas": 0,
            "erro": 0
        }

        # Filtro ativo (None = sem filtro)
        self.active_filter = None

        self.config = {
            "expected_files": ["CND MUNICIPAL", "CND RFB", "CND FGTS", "CND PROC", "CND ESTADUAL"],
            "target_line": "CERTIDÃO POSITIVA DE DÉBITOS - CPD",
            "last_folder": "",
            "mode": "Verificar Positiva",
            "ignored_folders": ["001 - RFB"]
        }

        self.load_config()
        self.create_dashboard()
        self.center_window()

        if self.config.get("last_folder") and os.path.exists(self.config["last_folder"]):
            self.folder_path.set(self.config["last_folder"])

        # Handler para fechamento seguro
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        logging.info("CND Dashboard iniciado")

    def on_closing(self):
        """Fecha a aplicação de forma segura"""
        self.is_closing = True
        self.cancel_requested = True

        # Parar executor se existir
        if self.executor:
            self.executor.shutdown(wait=False, cancel_futures=True)
            self.executor = None

        self.root.destroy()

    def safe_after(self, callback):
        """Executa callback na thread principal apenas se a janela ainda existir"""
        if self.is_closing:
            return
        try:
            self.root.after(0, callback)
        except Exception:
            pass

    def load_config(self):
        try:
            if os.path.exists("cnd_config.json"):
                with open("cnd_config.json", "r", encoding="utf-8") as f:
                    loaded_config = json.load(f)
                    self.config.update(loaded_config)
        except Exception as e:
            logging.error(f"Erro ao carregar configurações: {e}")

    def save_config(self):
        try:
            with open("cnd_config.json", "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            logging.error(f"Erro ao salvar configurações: {e}")

    def center_window(self):
        self.root.update_idletasks()
        width = self.root.winfo_width()
        height = self.root.winfo_height()
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def create_dashboard(self):
        # Container principal
        main_container = ctk.CTkFrame(self.root)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)

        # ============ HEADER ============
        header_frame = ctk.CTkFrame(main_container, height=60)
        header_frame.pack(fill="x", padx=5, pady=(5, 10))
        header_frame.pack_propagate(False)

        title_label = ctk.CTkLabel(header_frame, text="📊 CND Dashboard",
                                    font=ctk.CTkFont(size=24, weight="bold"))
        title_label.pack(side="left", padx=20, pady=10)

        # Controles no header
        controls_frame = ctk.CTkFrame(header_frame, fg_color="transparent")
        controls_frame.pack(side="right", padx=20, pady=10)

        self.mode_var = tk.StringVar(value=self.config["mode"])
        mode_combo = ctk.CTkComboBox(controls_frame, variable=self.mode_var,
                                      values=["Verificar Positiva", "Verificar Vencimento"],
                                      width=180, state="readonly")
        mode_combo.pack(side="left", padx=5)

        # ============ CONTENT AREA ============
        content_frame = ctk.CTkFrame(main_container, fg_color="transparent")
        content_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Lado esquerdo - Controles e Gráfico
        left_panel = ctk.CTkFrame(content_frame, width=350)
        left_panel.pack(side="left", fill="y", padx=(0, 10), pady=0)
        left_panel.pack_propagate(False)

        # Seleção de pasta
        folder_frame = ctk.CTkFrame(left_panel)
        folder_frame.pack(fill="x", padx=10, pady=10)

        folder_label = ctk.CTkLabel(folder_frame, text="📁 Pasta Principal:",
                                     font=ctk.CTkFont(size=12, weight="bold"))
        folder_label.pack(anchor="w", padx=10, pady=(10, 5))

        folder_input = ctk.CTkFrame(folder_frame, fg_color="transparent")
        folder_input.pack(fill="x", padx=10, pady=(0, 10))

        self.folder_entry = ctk.CTkEntry(folder_input, textvariable=self.folder_path,
                                          height=35, placeholder_text="Selecione uma pasta...")
        self.folder_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))

        browse_btn = ctk.CTkButton(folder_input, text="📂", width=40, height=35,
                                    command=self.browse_folder)
        browse_btn.pack(side="right")

        # Botões de ação
        action_frame = ctk.CTkFrame(left_panel)
        action_frame.pack(fill="x", padx=10, pady=5)

        # Frame para botões Processar e Parar lado a lado
        btn_row = ctk.CTkFrame(action_frame, fg_color="transparent")
        btn_row.pack(fill="x", padx=10, pady=10)

        self.process_btn = ctk.CTkButton(btn_row, text="▶ Processar",
                                          command=self.start_processing,
                                          height=40, font=ctk.CTkFont(size=13, weight="bold"),
                                          fg_color="#2d7d46", hover_color="#236b38")
        self.process_btn.pack(side="left", fill="x", expand=True, padx=(0, 5))

        self.stop_btn = ctk.CTkButton(btn_row, text="⏹ Parar",
                                       command=self.stop_processing,
                                       height=40, font=ctk.CTkFont(size=13, weight="bold"),
                                       fg_color="#b91c1c", hover_color="#991b1b",
                                       state="disabled")
        self.stop_btn.pack(side="right", fill="x", expand=True, padx=(5, 0))

        # Frame para Exportar e Limpar lado a lado
        btn_row2 = ctk.CTkFrame(action_frame, fg_color="transparent")
        btn_row2.pack(fill="x", padx=10, pady=(0, 10))

        self.export_btn = ctk.CTkButton(btn_row2, text="📥 Exportar",
                                         command=self.export_report,
                                         height=35, state="disabled",
                                         fg_color="#1a5fb4", hover_color="#144a8a")
        self.export_btn.pack(side="left", fill="x", expand=True, padx=(0, 5))

        self.clear_btn = ctk.CTkButton(btn_row2, text="🗑 Limpar",
                                        command=self.clear_data,
                                        height=35, state="disabled",
                                        fg_color="#6b7280", hover_color="#4b5563")
        self.clear_btn.pack(side="right", fill="x", expand=True, padx=(5, 0))

        # Progress
        self.progress_label = ctk.CTkLabel(action_frame, text="Aguardando...",
                                            font=ctk.CTkFont(size=11))
        self.progress_label.pack(pady=(5, 2))

        self.progress_bar = ctk.CTkProgressBar(action_frame, height=12)
        self.progress_bar.pack(fill="x", padx=10, pady=(0, 10))
        self.progress_bar.set(0)

        # ============ CARDS DE ESTATÍSTICAS ============
        stats_frame = ctk.CTkFrame(left_panel)
        stats_frame.pack(fill="x", padx=10, pady=10)

        stats_label = ctk.CTkLabel(stats_frame, text="📈 Estatísticas",
                                    font=ctk.CTkFont(size=14, weight="bold"))
        stats_label.pack(anchor="w", padx=10, pady=(10, 5))

        # Grid de cards
        cards_grid = ctk.CTkFrame(stats_frame, fg_color="transparent")
        cards_grid.pack(fill="x", padx=5, pady=5)

        # Criar cards (agora clicáveis com filter_key)
        # Linha 1: Total e Completas
        self.card_total = self.create_stat_card(cards_grid, "Total", "0", "#3b82f6", 0, 0, filter_key="total")
        self.card_completo = self.create_stat_card(cards_grid, "Completas", "0", "#22c55e", 0, 1, filter_key="completo")
        # Linha 2: Vencidas, Faltantes, Incompletas
        self.card_vencidas = self.create_stat_card(cards_grid, "Vencidas", "0", "#ef4444", 1, 0, filter_key="vencidas")
        self.card_faltantes = self.create_stat_card(cards_grid, "Faltantes", "0", "#f97316", 1, 1, filter_key="faltantes")

        # ============ GRÁFICO ============
        chart_frame = ctk.CTkFrame(left_panel)
        chart_frame.pack(fill="both", expand=True, padx=10, pady=10)

        chart_label = ctk.CTkLabel(chart_frame, text="📊 Distribuição",
                                    font=ctk.CTkFont(size=14, weight="bold"))
        chart_label.pack(anchor="w", padx=10, pady=(10, 5))

        self.chart_container = ctk.CTkFrame(chart_frame, fg_color="#1a1a2e")
        self.chart_container.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Placeholder para o gráfico
        self.chart_placeholder = ctk.CTkLabel(self.chart_container,
                                               text="Processe uma pasta\npara ver o gráfico",
                                               font=ctk.CTkFont(size=12),
                                               text_color="#666666")
        self.chart_placeholder.pack(expand=True)

        # ============ LADO DIREITO - TABELA ============
        right_panel = ctk.CTkFrame(content_frame)
        right_panel.pack(side="right", fill="both", expand=True)

        # Header da tabela com busca
        table_header = ctk.CTkFrame(right_panel, fg_color="transparent")
        table_header.pack(fill="x", padx=10, pady=10)

        table_title = ctk.CTkLabel(table_header, text="📋 Resultados",
                                    font=ctk.CTkFont(size=16, weight="bold"))
        table_title.pack(side="left")

        # Busca
        search_frame = ctk.CTkFrame(table_header, fg_color="transparent")
        search_frame.pack(side="right")

        self.search_entry = ctk.CTkEntry(search_frame, textvariable=self.search_var,
                                          width=250, height=32,
                                          placeholder_text="🔍 Buscar empresa...")
        self.search_entry.pack(side="left", padx=(0, 5))
        self.search_var.trace_add("write", self.filter_results)

        clear_btn = ctk.CTkButton(search_frame, text="✕", width=32, height=32,
                                   command=self.clear_search, fg_color="#4a4a4a")
        clear_btn.pack(side="left")

        # Tabela
        table_frame = ctk.CTkFrame(right_panel)
        table_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Estilo da tabela
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dashboard.Treeview",
                        background="#1a1a2e",
                        foreground="white",
                        fieldbackground="#1a1a2e",
                        rowheight=32,
                        font=('Segoe UI', 10))
        style.configure("Dashboard.Treeview.Heading",
                        background="#2d2d44",
                        foreground="white",
                        font=('Segoe UI', 10, 'bold'))
        style.map("Dashboard.Treeview",
                  background=[("selected", "#3b82f6")])

        # Criar Treeview
        self.tree = ttk.Treeview(table_frame, show="headings", height=15,
                                  style="Dashboard.Treeview")
        self.tree.pack(side="left", fill="both", expand=True)

        # Tags para cores (prioridade: vencida > faltando > incompleto > completo)
        self.tree.tag_configure("vencida", background="#6b1a1a", foreground="white")  # Vermelho escuro
        self.tree.tag_configure("faltando", background="#6b4a1a", foreground="white")  # Laranja escuro
        self.tree.tag_configure("incompleto", background="#4d4d1a", foreground="white")  # Amarelo escuro
        self.tree.tag_configure("erro", background="#4d1a4d", foreground="white")  # Roxo escuro
        self.tree.tag_configure("completo", background="#1a4d1a", foreground="white")  # Verde escuro
        self.tree.tag_configure("valida", background="#1a4d1a", foreground="white")  # Verde escuro

        # Scrollbar
        scrollbar = ctk.CTkScrollbar(table_frame, command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)

        # Bind para seleção
        self.tree.bind("<<TreeviewSelect>>", self.on_item_select)

        # ============ PAINEL DE DETALHES (bottom) ============
        self.details_frame = ctk.CTkFrame(right_panel, height=120)
        self.details_frame.pack(fill="x", padx=10, pady=(0, 10))
        self.details_frame.pack_propagate(False)

        details_label = ctk.CTkLabel(self.details_frame, text="📝 Detalhes da Empresa Selecionada",
                                      font=ctk.CTkFont(size=12, weight="bold"))
        details_label.pack(anchor="w", padx=15, pady=(10, 5))

        self.details_content = ctk.CTkLabel(self.details_frame,
                                             text="Selecione uma empresa na tabela para ver os detalhes",
                                             font=ctk.CTkFont(size=11),
                                             text_color="#888888",
                                             justify="left")
        self.details_content.pack(anchor="w", padx=15, pady=5)

    def create_stat_card(self, parent, title, value, color, row, col, filter_key=None):
        """Cria um card de estatística clicável"""
        card = ctk.CTkFrame(parent, fg_color="#2d2d44", corner_radius=10, cursor="hand2")
        card.grid(row=row, column=col, padx=5, pady=5, sticky="nsew")
        parent.grid_columnconfigure(col, weight=1)

        title_lbl = ctk.CTkLabel(card, text=title, font=ctk.CTkFont(size=11),
                                  text_color="#888888", cursor="hand2")
        title_lbl.pack(pady=(10, 2))

        value_lbl = ctk.CTkLabel(card, text=value, font=ctk.CTkFont(size=24, weight="bold"),
                                  text_color=color, cursor="hand2")
        value_lbl.pack(pady=(0, 10))

        # Tornar clicável
        if filter_key:
            card.bind("<Button-1>", lambda e, k=filter_key: self.filter_by_stat(k))
            title_lbl.bind("<Button-1>", lambda e, k=filter_key: self.filter_by_stat(k))
            value_lbl.bind("<Button-1>", lambda e, k=filter_key: self.filter_by_stat(k))

            # Efeito hover
            def on_enter(e, c=card):
                c.configure(fg_color="#3d3d54")
            def on_leave(e, c=card, k=filter_key):
                if self.active_filter == k:
                    c.configure(fg_color="#4a4a6a")
                else:
                    c.configure(fg_color="#2d2d44")

            card.bind("<Enter>", on_enter)
            card.bind("<Leave>", on_leave)
            title_lbl.bind("<Enter>", on_enter)
            title_lbl.bind("<Leave>", on_leave)
            value_lbl.bind("<Enter>", on_enter)
            value_lbl.bind("<Leave>", on_leave)

        return {"card": card, "value_lbl": value_lbl, "title_lbl": title_lbl}

    def filter_by_stat(self, filter_key):
        """Filtra a tabela pelo tipo de estatística clicado"""
        if not self.results_data:
            return

        # Se clicar no mesmo filtro, remove o filtro
        if self.active_filter == filter_key:
            self.active_filter = None
            self.filtered_data = self.results_data.copy()
        else:
            self.active_filter = filter_key
            mode = self.config["mode"]

            if filter_key == "total":
                self.filtered_data = self.results_data.copy()
                self.active_filter = None  # Total não é filtro real
            elif filter_key == "completo":
                self.filtered_data = [r for r in self.results_data
                                       if r.get("status", "").upper() == "COMPLETO"]
            elif filter_key == "incompleto":
                self.filtered_data = [r for r in self.results_data
                                       if r.get("status", "").upper() == "INCOMPLETO"]
            elif filter_key == "vencidas":
                if mode == "Verificar Vencimento":
                    self.filtered_data = [r for r in self.results_data
                                           if any(r.get(c) == "VENCIDA"
                                                  for c in ["municipal", "rfb", "fgts", "proc", "estadual"])]
                else:
                    # Modo positiva: filtra por erro
                    self.filtered_data = [r for r in self.results_data
                                           if r.get("status", "").upper() == "ERRO"]
            elif filter_key == "faltantes":
                # Filtra empresas que têm alguma CND faltando (NÃO)
                self.filtered_data = [r for r in self.results_data
                                       if any(r.get(c) == "NÃO"
                                              for c in ["municipal", "rfb", "fgts", "proc", "estadual"])]

        # Atualizar visual dos cards
        self.update_card_highlights()

        # Atualizar tabela
        for item in self.tree.get_children():
            self.tree.delete(item)

        for result in self.filtered_data:
            self.add_result_to_tree(result)

        # Atualizar label de filtro
        if self.active_filter:
            filter_names = {"completo": "Completas", "incompleto": "Incompletas", "vencidas": "Vencidas/Erros", "faltantes": "Faltantes"}
            self.progress_label.configure(text=f"Filtro: {filter_names.get(self.active_filter, '')} ({len(self.filtered_data)})")
        else:
            self.progress_label.configure(text=f"Total: {len(self.results_data)} empresas")

    def update_card_highlights(self):
        """Atualiza o destaque visual dos cards baseado no filtro ativo"""
        cards = {
            "total": self.card_total,
            "completo": self.card_completo,
            "vencidas": self.card_vencidas,
            "faltantes": self.card_faltantes
        }

        for key, card_data in cards.items():
            if self.active_filter == key:
                card_data["card"].configure(fg_color="#4a4a6a")
            else:
                card_data["card"].configure(fg_color="#2d2d44")

    def update_stats(self):
        """Atualiza os cards de estatísticas"""
        self.stats = {
            "total": len(self.results_data),
            "completo": 0,
            "incompleto": 0,
            "vencidas": 0,
            "faltantes": 0,
            "validas": 0,
            "erro": 0
        }

        mode = self.config["mode"]

        for r in self.results_data:
            status = r.get("status", "").upper()
            if status == "COMPLETO":
                self.stats["completo"] += 1
            elif status == "INCOMPLETO":
                self.stats["incompleto"] += 1
            elif status == "ERRO":
                self.stats["erro"] += 1

            # Verificar campos de CND
            campos = [r.get("municipal"), r.get("rfb"), r.get("fgts"),
                      r.get("proc"), r.get("estadual")]

            # Contar faltantes (NÃO em qualquer modo)
            if any(c == "NÃO" for c in campos):
                self.stats["faltantes"] += 1

            # Contar vencidas (só no modo vencimento)
            if mode == "Verificar Vencimento":
                if any(c == "VENCIDA" for c in campos):
                    self.stats["vencidas"] += 1
                elif all(c == "VÁLIDA" for c in campos if c and c != "NÃO"):
                    self.stats["validas"] += 1

        # Atualizar cards (agora usam dicionário)
        self.card_total["value_lbl"].configure(text=str(self.stats["total"]))
        self.card_completo["value_lbl"].configure(text=str(self.stats["completo"]))
        self.card_faltantes["value_lbl"].configure(text=str(self.stats["faltantes"]))

        if mode == "Verificar Vencimento":
            self.card_vencidas["value_lbl"].configure(text=str(self.stats["vencidas"]))
            self.card_vencidas["title_lbl"].configure(text="Vencidas")
        else:
            self.card_vencidas["value_lbl"].configure(text=str(self.stats["erro"]))
            self.card_vencidas["title_lbl"].configure(text="Erros")

        # Atualizar gráfico
        self.update_chart()

    def update_chart(self):
        """Atualiza o gráfico de pizza"""
        # Limpar container
        for widget in self.chart_container.winfo_children():
            widget.destroy()

        if not self.results_data:
            placeholder = ctk.CTkLabel(self.chart_container,
                                        text="Processe uma pasta\npara ver o gráfico",
                                        font=ctk.CTkFont(size=12),
                                        text_color="#666666")
            placeholder.pack(expand=True)
            return

        # Criar figura matplotlib
        fig, ax = plt.subplots(figsize=(3, 3), facecolor='#1a1a2e')
        ax.set_facecolor('#1a1a2e')

        # Dados para o gráfico
        labels = []
        sizes = []
        colors = []

        if self.stats["completo"] > 0:
            labels.append(f'Completas\n({self.stats["completo"]})')
            sizes.append(self.stats["completo"])
            colors.append('#22c55e')

        if self.stats["incompleto"] > 0:
            labels.append(f'Incompletas\n({self.stats["incompleto"]})')
            sizes.append(self.stats["incompleto"])
            colors.append('#ef4444')

        if self.stats["erro"] > 0:
            labels.append(f'Erros\n({self.stats["erro"]})')
            sizes.append(self.stats["erro"])
            colors.append('#eab308')

        if sizes:
            wedges, texts, autotexts = ax.pie(sizes, labels=labels, colors=colors,
                                               autopct='%1.0f%%', startangle=90,
                                               textprops={'color': 'white', 'fontsize': 8})
            for autotext in autotexts:
                autotext.set_fontsize(9)
                autotext.set_fontweight('bold')
        else:
            ax.text(0.5, 0.5, 'Sem dados', ha='center', va='center',
                    color='#666666', fontsize=12)

        ax.axis('equal')
        plt.tight_layout()

        # Adicionar ao tkinter
        canvas = FigureCanvasTkAgg(fig, master=self.chart_container)
        canvas.draw()
        canvas.get_tk_widget().pack(fill="both", expand=True)
        plt.close(fig)

    def on_item_select(self, event):
        """Mostra detalhes quando uma linha é selecionada"""
        selection = self.tree.selection()
        if not selection:
            return

        item = self.tree.item(selection[0])
        values = item["values"]

        if not values:
            return

        # Encontrar o resultado correspondente
        empresa = values[0]
        result = None
        for r in self.results_data:
            if r.get("empresa") == empresa:
                result = r
                break

        if result:
            mode = self.config["mode"]
            details = f"Empresa: {result.get('empresa', 'N/A')}\n"
            details += f"Status: {result.get('status', 'N/A')}\n"

            missing = result.get("missing_files", [])
            if missing:
                details += f"Arquivos faltantes: {', '.join(missing)}\n"

            outras = result.get("outras_cnds", [])
            if outras:
                details += f"Outras CNDs: {', '.join(outras[:3])}"
                if len(outras) > 3:
                    details += f" (+{len(outras)-3} mais)"

            self.details_content.configure(text=details, text_color="white")

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Selecione a pasta principal com as CNDs")
        if folder:
            self.folder_path.set(folder)
            self.config["last_folder"] = folder
            self.save_config()
            logging.info(f"Pasta selecionada: {folder}")

    def clear_search(self):
        self.search_var.set("")

    def clear_data(self):
        """Limpa todos os dados processados com confirmação"""
        if not self.results_data:
            return

        # Confirmação
        confirm = messagebox.askyesno(
            "Confirmar Limpeza",
            f"Deseja limpar todos os dados?\n\n"
            f"• {len(self.results_data)} empresas processadas serão removidas\n"
            f"• O gráfico será resetado\n"
            f"• Você poderá processar novamente",
            icon="warning"
        )

        if not confirm:
            return

        # Limpar dados
        self.results_data = []
        self.filtered_data = []
        self.active_filter = None

        # Limpar tabela
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Resetar estatísticas
        self.stats = {
            "total": 0,
            "completo": 0,
            "incompleto": 0,
            "vencidas": 0,
            "validas": 0,
            "erro": 0
        }

        # Atualizar cards
        self.card_total["value_lbl"].configure(text="0")
        self.card_completo["value_lbl"].configure(text="0")
        self.card_vencidas["value_lbl"].configure(text="0")
        self.card_faltantes["value_lbl"].configure(text="0")
        self.update_card_highlights()

        # Resetar gráfico
        self.update_chart()

        # Resetar progress
        self.progress_bar.set(0)
        self.progress_label.configure(text="Dados limpos - Pronto para processar")

        # Resetar detalhes
        self.details_content.configure(
            text="Selecione uma empresa na tabela para ver os detalhes",
            text_color="#888888"
        )

        # Desabilitar botões
        self.export_btn.configure(state="disabled")
        self.clear_btn.configure(state="disabled")

        logging.info("Dados limpos pelo usuário")

    def filter_results(self, *args):
        search_text = self.search_var.get().strip().lower()
        if not self.results_data:
            return

        for item in self.tree.get_children():
            self.tree.delete(item)

        if search_text:
            self.filtered_data = [r for r in self.results_data
                                   if search_text in r.get("empresa", "").lower()]
        else:
            self.filtered_data = self.results_data.copy()

        for result in self.filtered_data:
            self.add_result_to_tree(result)

    def sort_by_column(self, col):
        if not self.results_data:
            return

        col_map = {
            "Empresa": "empresa", "Municipal": "municipal", "RFB": "rfb",
            "FGTS": "fgts", "PROC": "proc", "Estadual": "estadual",
            "Positiva": "positiva", "Outras CNDs": "outras_cnds", "Status": "status"
        }

        key = col_map.get(col, "empresa")

        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False

        try:
            data_to_sort = self.filtered_data if self.filtered_data else self.results_data
            if key == "outras_cnds":
                sorted_data = sorted(data_to_sort, key=lambda x: len(x.get(key, [])),
                                      reverse=self.sort_reverse)
            else:
                sorted_data = sorted(data_to_sort, key=lambda x: str(x.get(key, "")).lower(),
                                      reverse=self.sort_reverse)

            for item in self.tree.get_children():
                self.tree.delete(item)

            for result in sorted_data:
                self.add_result_to_tree(result)

            direction = " ▼" if self.sort_reverse else " ▲"
            for c in self.tree["columns"]:
                text = c.replace(" ▲", "").replace(" ▼", "")
                if c == col or text == col:
                    self.tree.heading(c, text=text + direction)
                else:
                    self.tree.heading(c, text=text)

        except Exception as e:
            logging.error(f"Erro ao ordenar: {e}")

    def start_processing(self):
        if self.processing:
            return
        folder = self.folder_path.get().strip()
        if not folder or not os.path.exists(folder):
            messagebox.showerror("Erro", "Por favor, selecione uma pasta válida!")
            return

        self.config["mode"] = self.mode_var.get()
        self.save_config()

        self.processing = True
        self.cancel_requested = False
        self.process_btn.configure(text="⏳ Processando...", state="disabled")
        self.stop_btn.configure(state="normal")
        self.export_btn.configure(state="disabled")

        for item in self.tree.get_children():
            self.tree.delete(item)

        thread = threading.Thread(target=self.process_folder, args=(folder,))
        thread.daemon = True
        thread.start()

    def stop_processing(self):
        """Para o processamento em andamento"""
        if not self.processing:
            return
        self.cancel_requested = True
        self.stop_btn.configure(state="disabled", text="⏹ Parando...")
        logging.info("Cancelamento solicitado pelo usuário")

        # Cancelar executor se existir
        if self.executor:
            self.executor.shutdown(wait=False, cancel_futures=True)

    def process_folder(self, main_folder):
        try:
            start_time = time.time()
            logging.info(f"Iniciando processamento: {main_folder}")
            mode = self.config["mode"]
            expected_files = self.config["expected_files"]
            target_line = self.config["target_line"]

            # Filtrar pastas ignoradas
            ignored_folders = self.config.get("ignored_folders", [])
            subfolders = [f for f in os.listdir(main_folder)
                          if os.path.isdir(os.path.join(main_folder, f))
                          and f not in ignored_folders]
            total_folders = len(subfolders)

            if total_folders == 0:
                logging.warning("Nenhuma subpasta encontrada")
                self.safe_after(lambda: self.update_progress("Nenhuma subpasta!", 0))
                self.safe_after(self.processing_complete)
                return

            logging.info(f"Encontradas {total_folders} subpastas")
            self.results_data = []
            self.filtered_data = []

            # Usar ThreadPoolExecutor para processar em paralelo (otimização)
            max_workers = min(8, total_folders)  # Máximo 8 threads
            completed = 0

            def process_single(subfolder):
                # Verificar cancelamento ou fechamento antes de processar
                if self.cancel_requested or self.is_closing:
                    return None
                subfolder_path = os.path.join(main_folder, subfolder)
                if mode == "Verificar Positiva":
                    return self.process_subfolder_positive(subfolder_path, subfolder,
                                                           expected_files, target_line)
                else:
                    return self.process_subfolder_vencimento(subfolder_path, subfolder,
                                                             expected_files)

            self.executor = ThreadPoolExecutor(max_workers=max_workers)
            try:
                # Submeter todas as tarefas
                future_to_folder = {self.executor.submit(process_single, sf): sf for sf in subfolders}

                # Processar resultados conforme ficam prontos
                for future in as_completed(future_to_folder):
                    # Verificar cancelamento ou fechamento
                    if self.cancel_requested or self.is_closing:
                        logging.info("Processamento cancelado pelo usuário")
                        elapsed_time = time.time() - start_time
                        self.safe_after(lambda t=elapsed_time: self.update_progress(f"⏹ Cancelado em {t:.2f}s ({len(self.results_data)} processados)", 0))
                        break

                    completed += 1
                    progress = completed / total_folders
                    subfolder = future_to_folder[future]

                    try:
                        result = future.result()
                        if result is None:  # Foi cancelado
                            continue
                        self.results_data.append(result)
                        self.safe_after(lambda r=result: self.add_result_to_tree(r))
                        self.safe_after(lambda p=progress, f=subfolder:
                                        self.update_progress(f"Processando: {f} ({completed}/{total_folders})", p))
                    except Exception as e:
                        if not self.cancel_requested and not self.is_closing:
                            logging.error(f"Erro ao processar {subfolder}: {e}")
                            error_result = {"empresa": subfolder, "status": "ERRO", "outras_cnds": [], "missing_files": []}
                            self.results_data.append(error_result)
                            self.safe_after(lambda r=error_result: self.add_result_to_tree(r))
            finally:
                self.executor.shutdown(wait=False, cancel_futures=True)
                self.executor = None

            if not self.cancel_requested and not self.is_closing:
                elapsed_time = time.time() - start_time
                logging.info(f"Concluído: {len(self.results_data)} empresas em {elapsed_time:.2f}s (paralelo com {max_workers} threads)")
                self.safe_after(lambda t=elapsed_time: self.update_progress(f"✓ Concluído em {t:.2f}s", 1.0))

            self.safe_after(self.processing_complete)
            self.safe_after(self.update_stats)

        except Exception as e:
            logging.error(f"Erro: {e}", exc_info=True)
            if not self.is_closing:
                self.safe_after(lambda: messagebox.showerror("Erro", str(e)))
            self.safe_after(self.processing_complete)

    def process_subfolder_positive(self, subfolder_path, subfolder_name, expected_files, target_line):
        found_files = {file_type: False for file_type in expected_files}
        positive_cert_type = None
        outras_cnds = []
        try:
            for file_name in os.listdir(subfolder_path):
                if file_name.lower().endswith('.pdf'):
                    matched = False
                    for file_type in expected_files:
                        if file_type in file_name.upper():
                            matched = True
                            found_files[file_type] = True
                            pdf_path = os.path.join(subfolder_path, file_name)
                            if self.check_positive_cert(pdf_path, target_line):
                                positive_cert_type = file_type
                    if not matched:
                        outras_cnds.append(file_name)
            missing_files = [f for f, found in found_files.items() if not found]
            return {
                "empresa": subfolder_name,
                "municipal": "SIM" if found_files["CND MUNICIPAL"] else "NÃO",
                "rfb": "SIM" if found_files["CND RFB"] else "NÃO",
                "fgts": "SIM" if found_files["CND FGTS"] else "NÃO",
                "proc": "SIM" if found_files["CND PROC"] else "NÃO",
                "estadual": "SIM" if found_files["CND ESTADUAL"] else "NÃO",
                "positiva": positive_cert_type if positive_cert_type else "NENHUMA",
                "outras_cnds": outras_cnds,
                "status": "COMPLETO" if not missing_files else "INCOMPLETO",
                "missing_files": missing_files
            }
        except Exception as e:
            logging.error(f"Erro pasta '{subfolder_name}': {e}", exc_info=True)
            return {"empresa": subfolder_name, "status": "ERRO", "outras_cnds": [], "missing_files": []}

    def check_positive_cert(self, file_path, target_line):
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if page_text and target_line in page_text:
                        return True
            return False
        except Exception as e:
            logging.warning(f"Erro PDF '{file_path}': {e}")
            return False

    def check_due_date(self, file_name):
        try:
            match = re.search(r"\d{2}\.\d{2}\.\d{4}", file_name)
            if not match:
                return "DATA NÃO ENCONTRADA"
            date_str = match.group()
            due_date = datetime.strptime(date_str, "%d.%m.%Y").date()
            today = datetime.today().date()
            return "VENCIDA" if due_date < today else "VÁLIDA"
        except Exception as e:
            logging.warning(f"Erro data '{file_name}': {e}")
            return "ERRO DATA"

    def process_subfolder_vencimento(self, subfolder_path, subfolder_name, expected_files):
        found_files = {file_type: "NÃO" for file_type in expected_files}
        outras_cnds = []
        try:
            for file_name in os.listdir(subfolder_path):
                if file_name.lower().endswith('.pdf'):
                    matched = False
                    for file_type in expected_files:
                        if file_type in file_name.upper():
                            matched = True
                            found_files[file_type] = self.check_due_date(file_name)
                    if not matched:
                        status_venc = self.check_due_date(file_name)
                        outras_cnds.append(f"{file_name} ({status_venc})")
            missing_files = [f for f, status in found_files.items() if status == "NÃO"]
            return {
                "empresa": subfolder_name,
                "municipal": found_files["CND MUNICIPAL"],
                "rfb": found_files["CND RFB"],
                "fgts": found_files["CND FGTS"],
                "proc": found_files["CND PROC"],
                "estadual": found_files["CND ESTADUAL"],
                "outras_cnds": outras_cnds,
                "status": "COMPLETO" if not missing_files else "INCOMPLETO",
                "missing_files": missing_files
            }
        except Exception as e:
            logging.error(f"Erro pasta '{subfolder_name}': {e}", exc_info=True)
            return {"empresa": subfolder_name, "status": "ERRO", "outras_cnds": [], "missing_files": []}

    def update_progress(self, text, value):
        self.progress_label.configure(text=text)
        self.progress_bar.set(value)

    def add_result_to_tree(self, result):
        mode = self.config["mode"]
        outras = ", ".join(result.get("outras_cnds", [])[:2]) if result.get("outras_cnds") else "-"
        if len(result.get("outras_cnds", [])) > 2:
            outras += f" (+{len(result.get('outras_cnds', [])) - 2})"

        if mode == "Verificar Positiva":
            columns = ("Empresa", "Municipal", "RFB", "FGTS", "PROC", "Estadual", "Positiva", "Status")
            self.tree["columns"] = columns
            for col in columns:
                self.tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
                width = 150 if col == "Empresa" else 80
                self.tree.column(col, width=width, anchor="center")
            values = (result.get("empresa", ""), result.get("municipal", ""),
                      result.get("rfb", ""), result.get("fgts", ""),
                      result.get("proc", ""), result.get("estadual", ""),
                      result.get("positiva", ""), result.get("status", ""))
        else:
            columns = ("Empresa", "Municipal", "RFB", "FGTS", "PROC", "Estadual", "Status")
            self.tree["columns"] = columns
            for col in columns:
                self.tree.heading(col, text=col, command=lambda c=col: self.sort_by_column(c))
                width = 150 if col == "Empresa" else 90
                self.tree.column(col, width=width, anchor="center")
            values = (result.get("empresa", ""), result.get("municipal", ""),
                      result.get("rfb", ""), result.get("fgts", ""),
                      result.get("proc", ""), result.get("estadual", ""),
                      result.get("status", ""))

        # Determinar tag de cor com PRIORIDADE: vencida > faltando > incompleto > completo
        campos = [result.get("municipal"), result.get("rfb"), result.get("fgts"),
                  result.get("proc"), result.get("estadual")]
        status = result.get("status", "").upper()

        # Verificar condições em ordem de prioridade
        has_vencida = any(c == "VENCIDA" for c in campos)
        has_faltando = any(c == "NÃO" for c in campos)
        has_erro = status == "ERRO"

        # Determinar tag pela prioridade
        if has_vencida:
            tag = "vencida"  # Vermelho - mais crítico
        elif has_faltando:
            tag = "faltando"  # Laranja - CND não existe
        elif has_erro:
            tag = "erro"  # Roxo - erro no processamento
        elif status == "INCOMPLETO":
            tag = "incompleto"  # Amarelo
        elif status == "COMPLETO":
            tag = "completo"  # Verde
        else:
            tag = ""

        self.tree.insert("", "end", values=values, tags=(tag,) if tag else ())

    def processing_complete(self):
        self.processing = False
        self.cancel_requested = False
        self.process_btn.configure(text="▶ Processar", state="normal")
        self.stop_btn.configure(text="⏹ Parar", state="disabled")
        if self.results_data:
            self.export_btn.configure(state="normal")
            self.clear_btn.configure(state="normal")

    def export_report(self):
        if not self.results_data:
            messagebox.showwarning("Aviso", "Nenhum dado para exportar!")
            return
        try:
            filename = filedialog.asksaveasfilename(
                title="Salvar Relatório",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            if not filename:
                return
            self.create_excel_report(self.results_data, filename)
            messagebox.showinfo("Sucesso", f"Relatório exportado!\n{filename}")
        except Exception as e:
            messagebox.showerror("Erro", str(e))

    def create_excel_report(self, data, filename):
        wb = Workbook()
        ws = wb.active
        mode = self.config["mode"]
        ws.title = "Relatório CND"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style="thin"), right=Side(style="thin"),
                        top=Side(style="thin"), bottom=Side(style="thin"))

        if mode == "Verificar Positiva":
            headers = ["Empresa", "CND MUNICIPAL", "CND RFB", "CND FGTS", "CND PROC",
                       "CND ESTADUAL", "Certidão Positiva", "Outras CNDs", "Arquivos Faltantes", "Status"]
        else:
            headers = ["Empresa", "CND MUNICIPAL", "CND RFB", "CND FGTS", "CND PROC",
                       "CND ESTADUAL", "Outras CNDs", "Arquivos Faltantes", "Status"]

        # Título
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1,
                             value=f"RELATÓRIO DE CND ({mode}) - {datetime.now().strftime('%d/%m/%Y %H:%M')}")
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = center_alignment

        # Cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Dados
        for row_idx, result in enumerate(data, 3):
            outras = ", ".join(result.get("outras_cnds", [])) if result.get("outras_cnds") else "NENHUMA"
            if mode == "Verificar Positiva":
                row_data = [result["empresa"], result["municipal"], result["rfb"], result["fgts"],
                            result["proc"], result["estadual"], result.get("positiva", "NENHUMA"),
                            outras,
                            ", ".join(result["missing_files"]) if result["missing_files"] else "NENHUM",
                            result["status"]]
            else:
                row_data = [result["empresa"], result["municipal"], result["rfb"], result["fgts"],
                            result["proc"], result["estadual"], outras,
                            ", ".join(result["missing_files"]) if result["missing_files"] else "NENHUM",
                            result["status"]]

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = center_alignment

        # Formatação condicional
        opcoes = ["VENCIDA", "VÁLIDA", "IMPEDIDA", "TAREFA", "NÃO"]
        if mode == "Verificar Positiva":
            colunas_status = [2, 3, 4, 5, 6, 10]
        else:
            colunas_status = [2, 3, 4, 5, 6, 9]

        cores = {
            "VENCIDA": "FF0000", "VÁLIDA": "00FF00", "IMPEDIDA": "FFA500",
            "TAREFA": "0000FF", "NÃO": "808080", "COMPLETO": "00FF00",
            "INCOMPLETO": "FF0000"
        }

        for col in colunas_status:
            col_letter = chr(64 + col)
            dv = DataValidation(type="list", formula1=f'"{",".join(opcoes)}"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}3:{col_letter}500")

            for status, cor in cores.items():
                formula = f'EXACT("{status}",${col_letter}3)'
                rule = FormulaRule(formula=[formula],
                                   fill=PatternFill(start_color=cor, end_color=cor, fill_type="solid"))
                ws.conditional_formatting.add(f"{col_letter}3:{col_letter}500", rule)

        wb.save(filename)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = CNDDashboard()
    app.run()
